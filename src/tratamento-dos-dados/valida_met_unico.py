import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import argparse

# -------------------------------
# Utilidades de IO
# -------------------------------
def _ext(path: str) -> str:
    return os.path.splitext(path)[1].lower()

def _read_excel_any(path: str, header=None) -> pd.DataFrame:
    """
    Lê Excel em DataFrame:
      - .xls  -> engine='xlrd'   (xlrd==1.2.0)
      - .xlsx -> engine='openpyxl'
    """
    ext = _ext(path)
    if ext == ".xls":
        return pd.read_excel(path, header=header, engine="xlrd")
    elif ext == ".xlsx":
        return pd.read_excel(path, header=header, engine="openpyxl")
    else:
        raise ValueError(f"Extensão não suportada: {ext}")

def converter_xls_para_xlsx(filepath: str) -> str:
    """
    Se o arquivo for .xls, converte para .xlsx SEM usar pyexcel/Tk.
    Conversão: lê tudo via pandas (xlrd) e escreve um .xlsx (openpyxl).
    Mantém sem cabeçalho e sem index (compatível com o fluxo existente).
    """
    base, ext = os.path.splitext(filepath)
    if ext.lower() == ".xls":
        new_filepath = base + ".xlsx"
        df = _read_excel_any(filepath, header=None)
        # Escreve valores crus (sem header/index) para preservar grade
        with pd.ExcelWriter(new_filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, header=False)
        return new_filepath
    return filepath

# -------------------------------
# Lógica de detecção de versão
# -------------------------------
def detect_version(filepath: str):
    """
    Detecta a versão da planilha verificando o conteúdo das células:
      - Versão 3: se a célula B2 contém "EM11" (novo formato com coluna B).
      - Versão 1: se a célula C2 contém "EM11".
      - Versão 2: se a célula AA2 contém "EM11".
    Retorna 3, 1, 2 ou None se não encontrar.
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    try:
        if ws["B2"].value and "EM11" in str(ws["B2"].value):
            return 3
        if ws["C2"].value and "EM11" in str(ws["C2"].value):
            return 1
        if ws["AA2"].value and "EM11" in str(ws["AA2"].value):
            return 2
        return None
    finally:
        wb.close()

# -------------------------------
# Pipeline principal MET
# -------------------------------
def processar_met(filepath: str):
    """
    Processa met.xls/met.xlsx:
      - Se existir somente met.xlsx na pasta, cria cópia met_corrigido.xlsx e trabalha nela.
      - Se .xls, converte para .xlsx (pandas/xlrd -> openpyxl).
      - Detecta versão pelo cabeçalho e mapeia colunas.
      - Converte data para "YYYY-MM-DD HH:MM:SS".
      - Formata valores (ponto->vírgula; vazios->'n').
      - Gera met_corrigido.xlsx e met.csv (sem cabeçalho) na mesma pasta.
    """
    basename = os.path.basename(filepath).lower()
    dirpath = os.path.dirname(filepath)

    if basename not in ["met.xls", "met.xlsx"]:
        print(f"Arquivo {filepath} não possui o nome esperado 'met.xls' ou 'met.xlsx'.")
        return

    # Situação especial: apenas met.xlsx presente -> trabalhar sobre cópia met_corrigido.xlsx
    files_in_dir = [f.lower() for f in os.listdir(dirpath) if os.path.isfile(os.path.join(dirpath, f))]
    if files_in_dir.count("met.xlsx") == 1 and "met.xls" not in files_in_dir:
        corr_path = os.path.join(dirpath, "met_corrigido.xlsx")
        shutil.copy2(filepath, corr_path)
        filepath = corr_path
        print("Somente met.xlsx encontrado. Usando met_corrigido.xlsx como base para criação do CSV.")
    else:
        # Se veio .xls, converte
        filepath = converter_xls_para_xlsx(filepath)

    # Detecta versão
    version = detect_version(filepath)
    if not version:
        print(f"Formato diferente encontrado em {filepath}. Verifique a formatação da planilha.")
        return

    # Lê toda a planilha (dados começam na linha 9 – índice 8)
    df = _read_excel_any(filepath, header=None)

    # Escolhe mapeamento de colunas por versão
    if version == 1:
        col_data = 0
        col_vel  = 2
        col_dir  = 4
        col_prec = 6
        col_temp = 8
        col_umid = 12
        col_pres = 14
    elif version == 2:
        col_data = 0
        col_vel  = 26
        col_dir  = 28
        col_prec = 30
        col_temp = 32
        col_umid = 36
        col_pres = 38
    elif version == 3:
        col_data = 0
        col_vel  = 1
        col_dir  = 3
        col_prec = 5
        col_temp = 7
        col_umid = 11
        col_pres = 13
    else:
        print("Formato de planilha não reconhecido.")
        return

    # Seleciona dados a partir da linha 9
    df_measure = df.iloc[8:, [col_data, col_vel, col_dir, col_prec, col_temp, col_umid, col_pres]].copy()
    df_measure = df_measure[df_measure.iloc[:, 0].notna()]  # mantém apenas linhas com data

    # Data -> "YYYY-MM-DD HH:MM:SS"
    def parse_date(x):
        try:
            dt = pd.to_datetime(x, dayfirst=True, errors="raise")
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            print(f"Erro ao converter data '{x}': {e}")
            return "n"

    df_measure.iloc[:, 0] = df_measure.iloc[:, 0].apply(parse_date)

    # Demais valores -> 'n' se vazio; senão float->string com vírgula
    def format_val(x):
        if pd.isna(x):
            return "n"
        try:
            val = float(x)
            return str(val).replace(".", ",")
        except Exception:
            return str(x)

    for col_idx in range(1, 7):
        df_measure.iloc[:, col_idx] = df_measure.iloc[:, col_idx].apply(format_val)

    # Salva XLSX corrigido
    out_xlsx_path = os.path.join(dirpath, "met_corrigido.xlsx")
    df_measure.to_excel(out_xlsx_path, index=False, header=True)
    print(f"Arquivo XLSX criado/atualizado: {out_xlsx_path}")

    # Salva CSV sem cabeçalho
    out_csv_path = os.path.join(dirpath, "met.csv")
    df_measure.to_csv(out_csv_path, index=False, header=False, encoding="utf-8-sig")
    print(f"Arquivo CSV criado: {out_csv_path}")

# Execução via CLI (opcional)
if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Processa met.xls/met.xlsx e gera met_corrigido.xlsx + met.csv (headless).")
    ap.add_argument("arquivo", help="Caminho para met.xls ou met.xlsx")
    args = ap.parse_args()
    processar_met(args.arquivo)
