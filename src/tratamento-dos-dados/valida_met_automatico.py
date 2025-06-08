import os
import sys
import shutil
import pandas as pd
import pyexcel as p
from openpyxl import load_workbook
from datetime import datetime
from config import DADOS_COLETADOS_DIR

ROOT_DIR = DADOS_COLETADOS_DIR

def converter_xls_para_xlsx(filepath):
    """Converte arquivo .xls para .xlsx usando pyexcel, se necessário."""
    base, ext = os.path.splitext(filepath)
    if ext.lower() == '.xls':
        new_filepath = base + '.xlsx'
        p.save_book_as(file_name=filepath, dest_file_name=new_filepath)
        return new_filepath
    return filepath

def detect_version(filepath):
    """
    Detecta a versão da planilha a partir da presença do texto "EM11":
      - Versão 3: se a célula B2 contém "EM11" (novo formato com coluna B preenchida).
      - Versão 1: se a célula C2 contém "EM11".
      - Versão 2: se a célula AA2 contém "EM11".
    Retorna 3, 1, 2 ou None se não encontrar.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    if ws['B2'].value and "EM11" in str(ws['B2'].value):
        return 3
    if ws['C2'].value and "EM11" in str(ws['C2'].value):
        return 1
    if ws['AA2'].value and "EM11" in str(ws['AA2'].value):
        return 2
    return None

def processar_met(filepath):
    """
    Processa o arquivo met (met.xls ou met.xlsx) conforme:
      - Se na pasta existir somente met.xlsx, cria um met_corrigido.xlsx (cópia) e trabalha sobre ele.
      - Converte para XLSX se necessário.
      - Detecta a versão e extrai as colunas de interesse a partir da linha 9.
      - Converte a data para "YYYY-MM-DD HH:MM:SS" e formata os demais valores (troca ponto por vírgula e preenche vazios com "n").
      - Salva os resultados em met_corrigido.xlsx e met.csv (sem cabeçalho) na mesma pasta.
    """
    basename = os.path.basename(filepath).lower()
    dirpath = os.path.dirname(filepath)
    
    if basename not in ["met.xls", "met.xlsx"]:
        print(f"Arquivo {filepath} não possui o nome esperado 'met.xls' ou 'met.xlsx'.")
        return

    # Verifica quantos arquivos existem na pasta (em minúsculas)
    files_in_dir = [f.lower() for f in os.listdir(dirpath) if os.path.isfile(os.path.join(dirpath, f))]
    
    # Se existir somente met.xlsx na pasta (sem met.xls), cria um met_corrigido.xlsx a partir dele
    if files_in_dir.count("met.xlsx") == 1 and "met.xls" not in files_in_dir:
        corr_path = os.path.join(dirpath, "met_corrigido.xlsx")
        shutil.copy2(filepath, corr_path)
        filepath = corr_path
        print("Somente met.xlsx encontrado. Usando met_corrigido.xlsx como base para criação do CSV.")
    else:
        filepath = converter_xls_para_xlsx(filepath)
    
    version = detect_version(filepath)
    if not version:
        print(f"Formato diferente encontrado em {filepath}. Verifique a formatação da planilha.")
        return
    
    # Lê toda a planilha (os dados começam na linha 9 – índice 8)
    df = pd.read_excel(filepath, header=None, engine='openpyxl')
    
    # Define os índices das colunas conforme a versão detectada
    if version == 1:
        # Versão 1: coluna B está vazia, EM11 em C2
        col_data = 0   # Data em A
        col_vel  = 2   # Velocidade Escalar em C
        col_dir  = 4   # Direção Escalar em E
        col_prec = 6   # Precipitação em G
        col_temp = 8   # Temperatura em I
        col_umid = 12  # Umidade em M
        col_pres = 14  # Pressão em O
    elif version == 2:
        # Versão 2: dados deslocados (EM11 em AA2)
        col_data = 0    # Data em A
        col_vel  = 26   # Velocidade Escalar em AA
        col_dir  = 28   # Direção Escalar em AC
        col_prec = 30   # Precipitação em AE
        col_temp = 32   # Temperatura em AG
        col_umid = 36   # Umidade em AK
        col_pres = 38   # Pressão em AM
    elif version == 3:
        # Versão 3: novo formato com coluna B preenchida (EM11 de B até O)
        col_data = 0   # Data em A
        col_vel  = 1   # Velocidade Escalar em B
        col_dir  = 3   # Direção Escalar em D
        col_prec = 5   # Precipitação em F
        col_temp = 7   # Temperatura em H
        col_umid = 11  # Umidade em L
        col_pres = 13  # Pressão em N
    else:
        print("Formato de planilha não reconhecido.")
        return

    # Seleciona os dados a partir da linha 9 (índice 8)
    df_measure = df.iloc[8:, [col_data, col_vel, col_dir, col_prec, col_temp, col_umid, col_pres]].copy()
    df_measure = df_measure[df_measure.iloc[:, 0].notna()]  # Remove linhas sem data

    # Converte a coluna de data para o formato "YYYY-MM-DD HH:MM:SS"
    def parse_date(x):
        try:
            dt = pd.to_datetime(x, dayfirst=True, errors='raise')
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            print(f"Erro ao converter data '{x}': {e}")
            return "n"

    df_measure.iloc[:, 0] = df_measure.iloc[:, 0].apply(parse_date)

    # Formata os demais valores: se ausente, preenche com "n"; caso contrário, converte para float e troca ponto por vírgula
    def format_val(x):
        if pd.isna(x):
            return "n"
        try:
            val = float(x)
            return str(val).replace('.', ',')
        except:
            return str(x)

    for col_idx in range(1, 7):
        df_measure.iloc[:, col_idx] = df_measure.iloc[:, col_idx].apply(format_val)

    # Salva o arquivo XLSX corrigido (met_corrigido.xlsx) – sobrescrevendo, se já existir
    out_xlsx_path = os.path.join(dirpath, "met_corrigido.xlsx")
    df_measure.to_excel(out_xlsx_path, index=False, header=True)
    print(f"Arquivo XLSX criado/atualizado: {out_xlsx_path}")

    # Salva o arquivo CSV (sem cabeçalho)
    out_csv_path = os.path.join(dirpath, "met.csv")
    df_measure.to_csv(out_csv_path, index=False, header=False, encoding='utf-8-sig')
    print(f"Arquivo CSV criado: {out_csv_path}")

def processar_diretorios(root_dir):
    """
    Percorre todas as pastas de ano e mês dentro de root_dir.
    Processa os arquivos 'met.xls' ou 'met.xlsx' em cada pasta de mês,
    respeitando que, para o ano corrente, somente os meses já ocorridos serão processados.
    """
    now = datetime.now()
    current_year = now.year
    current_month = now.month

    # Mapeia nomes de meses em português para número
    months_map = {
        'janeiro': 1,
        'fevereiro': 2,
        'março': 3,
        'abril': 4,
        'maio': 5,
        'junho': 6,
        'julho': 7,
        'agosto': 8,
        'setembro': 9,
        'outubro': 10,
        'novembro': 11,
        'dezembro': 12,
    }

    # Percorre as pastas de ano
    for year_folder in sorted(os.listdir(root_dir)):
        year_path = os.path.join(root_dir, year_folder)
        if not os.path.isdir(year_path):
            continue
        try:
            year = int(year_folder)
        except ValueError:
            print(f"Pasta ignorada (não é ano): {year_folder}")
            continue
        if year > current_year:
            print(f"Pasta {year_path} ignorada por ser de um ano futuro.")
            continue

        # Percorre as pastas de mês dentro do ano
        for month_folder in sorted(os.listdir(year_path)):
            month_path = os.path.join(year_path, month_folder)
            if not os.path.isdir(month_path):
                continue
            month_name = month_folder.lower()
            if month_name not in months_map:
                print(f"Pasta {month_path} ignorada: nome de mês não reconhecido.")
                continue
            month_number = months_map[month_name]
            if year == current_year and month_number > current_month:
                print(f"Pasta {month_path} ignorada: mês futuro.")
                continue

            # Procura por met.xls ou met.xlsx na pasta do mês
            met_file = None
            for candidate in ["met.xls", "met.xlsx"]:
                candidate_path = os.path.join(month_path, candidate)
                if os.path.isfile(candidate_path):
                    met_file = candidate_path
                    break
            if met_file is None:
                print(f"Arquivo 'met' não encontrado em {month_path}.")
            else:
                print(f"Processando: {met_file}")
                processar_met(met_file)

    print("Processamento automático concluído.")
    sys.exit()

if __name__ == "__main__":
    processar_diretorios(ROOT_DIR)
