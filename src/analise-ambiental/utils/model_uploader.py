# utils/model_uploader.py
# ------------------------------------------------------------
# Uploader/validador sem dependência de Tk:
# - Aceita "qar.xls/.xlsx" ou "met.xls/.xlsx"
# - Lê a célula A100 via xlrd (xls) ou openpyxl (xlsx)
# - A100 deve conter data/hora; infere <ano>/<mês>
# - Bloqueia se já existir arquivo do mesmo tipo no destino
# - Executa valida_qar_unico / valida_met_unico e grava APENAS o CSV final
#   em dados-coletados/<ano>/<mês>/qar.csv (ou met.csv)
# - Em caso de DUPLICATE, retorna tipo/ano/mês p/ permitir o MERGE
# - Em caso de erro, também retorna tipo/ano/mês/dest_dir quando possível
# ------------------------------------------------------------

from __future__ import annotations

import os
import io
import sys
import time
import glob
import shutil
import traceback
from types import ModuleType
from typing import Dict, Optional, Tuple
from datetime import datetime, timedelta

from werkzeug.datastructures import FileStorage

# Caminho raiz do projeto (…/analise-ambiental/utils -> sobe 2 níveis)
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

# Configs (caminhos principais)
try:
    from config import DATA_ROOT, UPLOAD_TMP  # preferir config do app, se existir
except Exception:
    # Fallbacks sensatos caso não exista config.py (para testes isolados)
    DATA_ROOT = os.path.join(ROOT_DIR, "dados-coletados")
    UPLOAD_TMP = os.path.join(ROOT_DIR, "tmp-uploads")

# ------------------------------------------------------------
# Meses em PT usados como nome de pasta
# ------------------------------------------------------------
PT_MONTHS = [
    "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
]

def _month_name_pt(n: int) -> str:
    if not (1 <= n <= 12):
        raise ValueError(f"Mês inválido: {n}")
    return PT_MONTHS[n - 1]

# ------------------------------------------------------------
# Utils de FS
# ------------------------------------------------------------
def _ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def _make_job_dir() -> str:
    _ensure_dir(UPLOAD_TMP)
    d = os.path.join(UPLOAD_TMP, f"job_{int(time.time()*1000)}")
    _ensure_dir(d)
    return d

def _canonical_dest_dir(year: int, month: int) -> str:
    return os.path.join(DATA_ROOT, str(year), _month_name_pt(month))

def _ci_in_dir(path: str, names: Tuple[str, ...]) -> list[str]:
    """Case-insensitive lookup de nomes dentro da pasta."""
    if not os.path.isdir(path):
        return []
    present = []
    lower = {f.lower(): f for f in os.listdir(path)}
    for n in names:
        if n.lower() in lower:
            present.append(lower[n.lower()])
    return present

def _has_any_of(path: str, candidates: Tuple[str, ...]) -> list[str]:
    """Verifica se qualquer um dos arquivos existe (case-insensitive)."""
    found = []
    for name in candidates:
        found.extend(_ci_in_dir(path, (name,)))
    return found

# ------------------------------------------------------------
# Leitura de A100 usando xlrd (xls) ou openpyxl (xlsx) — SEM TK
# ------------------------------------------------------------
def _parse_datetime_like(val) -> datetime:
    """Converte valores comuns (datetime, string, número serial Excel) em datetime."""
    # 1) datetime direto
    if isinstance(val, datetime):
        return val

    # 2) string "m/d/yyyy HH:MM" (e variações)
    if isinstance(val, str):
        s = val.strip()
        fmts = [
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%y %H:%M",
            "%m/%d/%y %H:%M:%S",
            "%m/%d/%Y",
            "%m/%d/%y",
        ]
        for f in fmts:
            try:
                return datetime.strptime(s, f)
            except Exception:
                pass
        raise ValueError(f"A100 não está em um formato de data/hora aceito: '{s}'")

    # 3) número serial Excel (base 1899-12-30)
    if isinstance(val, (int, float)):
        base = datetime(1899, 12, 30)
        return base + timedelta(days=float(val))

    raise ValueError(f"Valor inesperado em A100: {type(val).__name__}: {val!r}")

def _read_a100_datetime(file_path: str) -> datetime:
    """
    Lê A100 (linha 100, coluna A) do primeiro sheet.
    - Para .xls usa xlrd (requer xlrd<=1.2.0 para .xls)
    - Para .xlsx usa openpyxl
    Retorna datetime; lança ValueError se não conseguir.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xls":
        try:
            import xlrd  # xlrd 1.2.0 suporta .xls
        except Exception as e:
            raise ValueError(
                "xlrd não está disponível. Instale 'xlrd==1.2.0' para ler .xls."
            ) from e

        try:
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)
            val = sheet.cell_value(99, 0)  # A100
            # Se for float, converter com datemode (melhor que base fixa)
            if isinstance(val, (int, float)):
                try:
                    val = xlrd.xldate_as_datetime(val, book.datemode)
                except Exception:
                    # fallback: trata como serial puro
                    pass
            return _parse_datetime_like(val)
        except Exception as e:
            raise ValueError(f"Não foi possível ler a célula A100 (.xls): {e}")

    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise ValueError(
                "openpyxl não está disponível. Instale 'openpyxl' para ler .xlsx."
            ) from e

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            val = ws["A100"].value
            wb.close()
            return _parse_datetime_like(val)
        except Exception as e:
            raise ValueError(f"Não foi possível ler a célula A100 (.xlsx): {e}")

    else:
        # Pelo fluxo normal só aceitamos .xls/.xlsx
        raise ValueError(f"Extensão não suportada para leitura de A100: '{ext}'. Use .xls ou .xlsx.")

# ------------------------------------------------------------
# Carrega módulos de validação (tratamento-dos-dados/*.py)
# ------------------------------------------------------------
def _load_module_from_path(mod_name: str, file_path: str) -> ModuleType:
    import importlib.util
    spec = importlib.util.spec_from_file_location(mod_name, file_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Não foi possível carregar o módulo em {file_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)  # type: ignore
    return module

def _run_qar_validator(src_xls: str, work_dir: str) -> str:
    """
    Executa valida_qar_unico.processar_qar(src_xls) no work_dir.
    Retorna o caminho do CSV gerado (qar.csv, qar_maior.csv ou qar_novo.csv).
    """
    val_path = os.path.join(ROOT_DIR, "tratamento-dos-dados", "valida_qar_unico.py")
    mod = _load_module_from_path("valida_qar_unico", val_path)

    if not hasattr(mod, "processar_qar"):
        raise RuntimeError("valida_qar_unico.py não expõe a função processar_qar(filepath).")

    # os validadores assumem cwd? garanta que outputs caiam no work_dir
    old_cwd = os.getcwd()
    try:
        os.chdir(work_dir)
        mod.processar_qar(src_xls)
    finally:
        os.chdir(old_cwd)

    candidates = [
        os.path.join(work_dir, "qar.csv"),
        os.path.join(work_dir, "qar_maior.csv"),
        os.path.join(work_dir, "qar_novo.csv"),
    ]
    found = [c for c in candidates if os.path.isfile(c)]
    if not found:
        globbed = sorted(glob.glob(os.path.join(work_dir, "qar*.csv")), key=os.path.getmtime, reverse=True)
        if globbed:
            return globbed[0]
        raise RuntimeError("Validador QAR executou, mas o CSV final não foi encontrado.")
    found.sort(key=os.path.getmtime, reverse=True)
    return found[0]

def _run_met_validator(src_xls: str, work_dir: str) -> str:
    """
    Executa valida_met_unico.processar_met(src_xls) no work_dir.
    Retorna o caminho do CSV gerado (sempre 'met.csv').
    """
    val_path = os.path.join(ROOT_DIR, "tratamento-dos-dados", "valida_met_unico.py")
    mod = _load_module_from_path("valida_met_unico", val_path)

    if not hasattr(mod, "processar_met"):
        raise RuntimeError("valida_met_unico.py não expõe a função processar_met(filepath).")

    old_cwd = os.getcwd()
    try:
        os.chdir(work_dir)
        mod.processar_met(src_xls)
    finally:
        os.chdir(old_cwd)

    out_csv = os.path.join(work_dir, "met.csv")
    if not os.path.isfile(out_csv):
        raise RuntimeError("Validador MET executou, mas 'met.csv' não foi encontrado.")
    return out_csv

# ------------------------------------------------------------
# Regras de duplicidade
# ------------------------------------------------------------
def _check_duplicate(tipo: str, dest_dir: str) -> Optional[Dict]:
    """
    Bloqueia se já existir arquivo do mesmo tipo no destino.
    Considera: .xls, .xlsx, .csv daquele tipo.
    """
    if tipo == "qar":
        candidates = ("qar.xls", "qar.xlsx", "qar.csv")
    elif tipo == "met":
        candidates = ("met.xls", "met.xlsx", "met.csv")
    else:
        return {"ok": False, "code": "invalid_type", "message": f"Tipo inválido '{tipo}'."}

    found = _has_any_of(dest_dir, candidates)
    if found:
        return {
            "ok": False,
            "code": "duplicate",
            "message": f"Já existe arquivo de {tipo.upper()} nesta pasta ({dest_dir}): {', '.join(found)}. Não foi aceito.",
            "dest_dir": dest_dir,
            "existing": found,
        }
    return None

# ------------------------------------------------------------
# Função principal chamada pela API
# ------------------------------------------------------------
def handle_uploaded_model(file: FileStorage) -> Dict:
    """
    Parâmetro: FileStorage (campo 'file' do FormData).
    Passos:
      1) Verifica nome: 'qar.xls/.xlsx' ou 'met.xls/.xlsx'
      2) Lê A100, infere ano/mês (xlrd/openpyxl — sem Tk)
      3) Bloqueia se duplicado (tipo no destino)
      4) Executa validador correspondente em pasta temporária
      5) Copia CSV final para dados-coletados/<ano>/<mês> como 'qar.csv' ou 'met.csv'
      6) NÃO salva o .xls original no dados-coletados
    Retorna um dicionário JSON-friendly.
    """
    if not file or not getattr(file, "filename", None):
        return {"ok": False, "code": "no_file", "message": "Nenhum arquivo enviado."}

    original_name = (file.filename or "").strip()
    name_lower = original_name.lower()

    # Aceitar .xlsx também
    allowed = ("qar.xls", "met.xls", "qar.xlsx", "met.xlsx")
    if name_lower not in allowed:
        return {
            "ok": False,
            "code": "invalid_name",
            "message": "O arquivo deve se chamar exatamente 'qar.xls'/'qar.xlsx' ou 'met.xls'/'met.xlsx'. Renomeie e envie novamente."
        }

    tipo = "qar" if name_lower.startswith("qar") else "met"

    # Área temporária deste job
    job_dir = _make_job_dir()
    tmp_xls_path = os.path.join(job_dir, name_lower)  # nome exato esperado pelos validadores
    file.save(tmp_xls_path)

    year = None
    month = None
    dest_dir = None

    try:
        # 1) lê A100 para decidir ano/mês (sem Tk)
        dt = _read_a100_datetime(tmp_xls_path)
        year = dt.year
        month = dt.month

        dest_dir = _canonical_dest_dir(year, month)

        # 2) checa duplicidade (resposta enriquecida com tipo/ano/mês)
        dup = _check_duplicate(tipo, dest_dir)
        if dup:
            dup.update({"tipo": tipo, "year": year, "month": _month_name_pt(month)})
            return dup

        # 3) roda validador e copia CSV final
        _ensure_dir(dest_dir)

        if tipo == "qar":
            out_csv_tmp = _run_qar_validator(tmp_xls_path, job_dir)
            out_final = os.path.join(dest_dir, "qar.csv")
        else:
            out_csv_tmp = _run_met_validator(tmp_xls_path, job_dir)
            out_final = os.path.join(dest_dir, "met.csv")

        if os.path.exists(out_final):
            raise FileExistsError(f"Arquivo destino já existe: {out_final}")

        shutil.copy2(out_csv_tmp, out_final)

        return {
            "ok": True,
            "tipo": tipo,
            "year": year,
            "month": _month_name_pt(month),
            "dest_dir": dest_dir,
            "saved": True,
            "message": f"✔ {tipo.upper()} validado e salvo em '{out_final}'."
        }

    except Exception as e:
        tb = traceback.format_exc(limit=2)
        # Retornar o máximo de contexto possível para permitir merge a jusante
        payload = {
            "ok": False,
            "code": "validator_error",
            "message": f"Erro ao validar e gerar CSV de {str(tipo).upper() if isinstance(tipo, str) else tipo}: {e}",
            "hint": "Verifique se a planilha segue o modelo. A100 deve conter a data no formato mês/dia/ano hh:mm.",
            "trace": tb,
        }
        if tipo: payload["tipo"] = tipo
        if year: payload["year"] = year
        if month: payload["month"] = _month_name_pt(month)
        if dest_dir: payload["dest_dir"] = dest_dir
        return payload
    finally:
        try:
            shutil.rmtree(job_dir, ignore_errors=True)
        except Exception:
            pass
